import os
import openai
import telebot
from openpyxl import load_workbook
import datetime
from dotenv import load_dotenv, find_dotenv

wb_name = 'message_list.xlsx'

workbook = load_workbook(wb_name)
worksheet = workbook['Sheet1']


def set_actual_count():
    count = 2
    value = 1
    while value != None:
        cell = worksheet.cell(row=count, column=1)
        value = cell.value
        count += 1
    return count - 1


load_dotenv(find_dotenv())
count = set_actual_count()
openai.api_key = os.getenv('OPENAI_KEY')
bot = telebot.TeleBot(os.getenv('TELEGRAM_KEY'))


@bot.message_handler(func=lambda _: True)
def handler_message(message):
    global count
    response = openai.Completion.create(
        model="text-davinci-003",
        prompt=message.text,
        temperature=0.5,
        max_tokens=1000,
        top_p=1.0,
        frequency_penalty=0.5,
        presence_penalty=0.0,

    )
    if message.text == '/start':

        bot.send_message(chat_id=message.from_user.id,
                         text=f' Привет {message.from_user.first_name}! Задай мне свой вопрос!')
    else:
        bot.send_message(chat_id=message.from_user.id,
                         text=f'{message.from_user.first_name}, {response["choices"][0]["text"]}')

    worksheet[f'A{count}'] = str(datetime.datetime.now().date())
    worksheet[f'B{count}'] = str(datetime.datetime.now().time())[0:8]
    worksheet[f'C{count}'] = f'{message.from_user.first_name} {message.from_user.last_name}'
    worksheet[f'D{count}'] = message.text
    worksheet[f'E{count}'] = message.from_user.id
    worksheet[f'F{count}'] = response["choices"][0]["text"]

    count += 1
    workbook.save(wb_name)


try:
    bot.polling()
except:
    workbook.save(wb_name)
